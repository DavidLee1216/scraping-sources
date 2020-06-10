from bs4 import BeautifulSoup
from selenium import webdriver
import time as ttime
import xlsxwriter
import xlrd
import openpyxl
import sys
import logging
import requests
from urllib import request
import urllib
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# driver = webdriver.Chrome(executable_path='./chromedrive/chromedriver.exe')
# driver.set_page_load_timeout(50)

user_agent = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/63.0.3239.84 Chrome/63.0.3239.84 Safari/537.36'

header = {'user-agent': user_agent}

python_version = sys.version_info.major

if python_version == 3:
    import urllib.parse
    import urllib.request
    urljoin = urllib.parse.urljoin
    urlretrieve = urllib.request.urlretrieve
    quote = urllib.parse.quote

    # configure headers
    opener = urllib.request.build_opener()
    opener.addheaders = [('User-agent', user_agent)]
    urllib.request.install_opener(opener)

line = 1
workbook = None
worksheet = None

def analyse():
    global line

    url = 'https://fileinfo.com/browse/'

    wb = openpyxl.load_workbook('fileinfo.xlsx')
    idx1 = wb.sheetnames.index('Sheet1')
    wb.remove(wb.worksheets[idx1])
    wb.create_sheet('Sheet1')
    sheet1 = wb.active
    sheet1.title = 'Sheet1'
    sheet1['A1'] = 'SI#'
    sheet1['B1'] = 'File Extension'
    sheet1['C1'] = 'Flie Type'
    sheet1['D1'] = 'category'
    sheet1['E1'] = 'Developer'
    sheet1['F1'] = 'Format'
    sheet1['G1'] = 'Windows'
    sheet1['H1'] = 'Mac'
    sheet1['I1'] = 'Linux'
    sheet1['J1'] = 'iOS'
    sheet1['K1'] = 'Android'
    sheet1['L1'] = 'Web'

#    try:
#        driver.maximize_window()
#        driver.get(url)
#        ttime.sleep(2)
    # except:
    #     try:
    #         driver.get(url)
    #         ttime.sleep(2)
    #     except:
    #         return -1
    try:
        content = requests.get(url, headers=header).content
        # driver.get(url)
        # ttime.sleep(2)
        # content = driver.page_source
        soup = BeautifulSoup(content, "html.parser")
        ii = 0
        for link in soup.find('div', class_='alpha').find_all('a'):
            # if ii == 0:
            #     ii += 1
            #     continue
            link_url = link.get('href')
            file_list_url = urljoin(url, link_url)
            # driver.get(file_list_url)
            # ttime.sleep(2)
            # file_list_content = driver.page_source
            file_list_content = requests.get(file_list_url, headers=header).content
            sub_list_soup = BeautifulSoup(file_list_content, 'html.parser')
            for sub_link in sub_list_soup.find('table', class_='slist').find('tbody').find_all('tr'):
                sub_link_url = sub_link.find('a').get('href')
                file_content_url = urljoin(url, sub_link_url)
                # driver.get(file_content_url)
                # ttime.sleep(5)
                # file_content = driver.page_source
                file_content = requests.get(file_content_url, headers=header).content
                file_content_soup = BeautifulSoup(file_content, 'html.parser')
                extension = file_content_soup.find('article').find('h1').find('b').get_text('|', strip=True)
                sections = file_content_soup.find_all('section', class_='ext')
                for section in sections:
                    file_type_item = section.find('h2')
                    file_type = file_type_item.find('span').next_sibling.strip()
                    file_header = section.find('div', class_='fileHeader').find('table', class_='headerInfo')
                    header_info = file_header.find_all('tr')
                    developer = header_info[0].find_all('td')[1].get_text('|', strip=True)
                    category = header_info[2].find_all('td')[1].get_text('|', strip=True)
                    formatt = header_info[3].find_all('td')[1].get_text('|', strip=True)
                    windows = ''
                    mac = ''
                    linux = ''
                    ios = ''
                    android = ''
                    web = ''
                    programs_table = section.find_all('table', class_='programs')
                    
                    for program in programs_table:
                        platform = program.find('td', class_='platform').get_text('|', strip=True)
                        apps = program.find('table', class_='apps').find_all('tr', recursive=False)
                        app_string = ''
                        for app in apps:
                            ss = app.get_text('|', strip=True)
                            if app_string != '':
                                app_string += ','
                            app_string += ss
                        if platform == 'Windows':
                            windows = app_string
                        elif platform == 'Mac':
                            mac = app_string
                        elif platform == 'Linux':
                            linux = app_string
                        elif platform == 'iOS':
                            ios = app_string
                        elif platform == 'Android':
                            android = app_string
                        elif platform == 'Web':
                            web = app_string
                    sheet1.append([line, extension, file_type, category, developer, formatt, windows, mac, linux, ios, android, web])
                    logging.warning("ext: %s, line: %d", extension, line)
                    line += 1
                wb.save('fileinfo.xlsx')

    except Exception as e:
        logging.warning("exception: %s" % e)
        

if __name__ == "__main__":
    analyse()