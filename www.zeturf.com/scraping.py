from bs4 import BeautifulSoup
import sys
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import datetime
import time as ttime
import xlsxwriter
import xlrd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

#options = Options()
#options.add_experimental_option('excludeSwitches', ['enable-logging'])
#driver = webdriver.Chrome(executable_path='./chromedrive/chromedriver.exe', chrome_options=options)

user_agent = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/63.0.3239.84 Chrome/63.0.3239.84 Safari/537.36'

python_version = sys.version_info.major

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
DAYS = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

if python_version == 3:
    import urllib.parse
    import urllib.request
    urljoin = urllib.parse.urljoin
    urlretrieve = urllib.request.urlretrieve
    quote = urllib.parse.quote

    # configure headers
    opener = urllib.request.build_opener()
    opener.addheaders = [('User-agent', user_agent)]
    urllib.request.install_opener(opener)

line = 2

def make_excel_header(worksheet):
    worksheet.write("A1", "Date")
    worksheet.write("B1", "Country")
    worksheet.write("C1", "Racetrack")
    worksheet.write("D1", "Race name ")
    worksheet.write("E1", "Race type")
    worksheet.write("F1", "Race link")
    worksheet.write("G1", "Conditions of the race")
    worksheet.write("H1", "Starters")
    worksheet.write("I1", "Distance")
    worksheet.write("J1", "Winner")
    worksheet.write("K1", "Arrivees")
    worksheet.write("L1", "Runners")
    worksheet.write("M1", "Name")
    worksheet.write("N1", "Sex")
    worksheet.write("O1", "Age")
    worksheet.write("P1", "Distance")
    worksheet.write("Q1", "Driver")
    worksheet.write("R1", "Trainer")
    worksheet.write("S1", "Musique")
    worksheet.write("T1", "Odds")
    worksheet.write("U1", "Time")
    worksheet.write("V1", "Racing time")
    worksheet.write("W1", "Position")

def make_excel_header_openpy(worksheet):
    worksheet["A1"] = "Date"
    worksheet["B1"] = "Country"
    worksheet["C1"] = "Racetrack"
    worksheet["D1"] = "Race name "
    worksheet["E1"] = "Race type"
    worksheet["F1"] = "Race link"
    worksheet["G1"] = "Conditions of the race"
    worksheet["H1"] = "Gagne limit"
    worksheet["I1"] = "Starters"
    worksheet["J1"] = "Distance"
    worksheet["K1"] = "Winner"
    worksheet["L1"] = "Arrivees"
    worksheet["M1"] = "Runners"
    worksheet["N1"] = "Name"
    worksheet["O1"] = "Info"
    worksheet["P1"] = "Sex"
    worksheet["Q1"] = "Age"
    worksheet["R1"] = "Distance"
    worksheet["S1"] = "Driver"
    worksheet["T1"] = "Trainer"
    worksheet["U1"] = "Musique"
    worksheet["V1"] = "Odds"
    worksheet["W1"] = "Time"
    worksheet["X1"] = "Racing time"
    worksheet["Y1"] = "Position"

def open_page_for_date(driver, url, year, month, date):
    validate_button = driver.find_element_by_xpath("//button[@id='filre-archives']")
    year_option_menu = driver.find_element_by_id('year').find_elements_by_tag_name('option')
    for y in year_option_menu:
        if int(y.text)==year:
            y.click()
            month_option_menu = driver.find_element_by_id('month').find_elements_by_tag_name('option')
            for m in month_option_menu:
                if MONTHS.index(m.text.strip())==month-1:
                    m.click()
                    driver.execute_script("arguments[0].click();", validate_button)
                    ttime.sleep(2)
                    break
            break
    soup = BeautifulSoup(driver.page_source, "html.parser")
    months = soup.find('div', class_='calendar').find_all('section', class_='month')
    for month_item in months:
        month_str = month_item.find('span', class_='header').get_text('|', strip=True)
        if month_str != MONTHS[month-1]:
            continue
        dates = month_item.find('div', class_='dates').find_all('span', class_='active')
        for date_item in dates:
            date_str = date_item.get_text('|', strip=True)
            if int(date_str) == date:
                link_url = date_item.find('a').get('href')
                link_url = urljoin(url, link_url)
                return link_url
    return None
    
def make_date_url(year, month, date):
    url = "https://www.zeturf.com/en/resultats-et-rapports/"
    string = "{:d}-{:02d}-{:02d}/turf".format(year, month, date)
    return url+string

class RacingTrackInfo:
    def __init__(self, country, race_track, url):
        self.country = country
        self.race_track = race_track
        self.url = url

class RacingGeneralInfo:
    def __init__(self, racing_name, racing_type, starters, distance, winner, runners, url):
        self.racing_name = racing_name
        self.racing_type = racing_type
        self.starters = starters
        self.distance = distance
        self.winner = winner
        self.runners = runners
        self.url = url

class RacingRunnerInfo:
    def __init__(self, runner_number, runner_name, info, sex, age, distance, driver, trainer, musique, odds, time_val, racing_time, position):
        self.runner_number = runner_number
        self.runner_name = runner_name
        self.info = info
        self.sex = sex
        self.age = age
        self.distance = distance
        self.driver = driver
        self.traniner = trainer
        self.musique = musique
        self.odds = odds
        self.time_val = time_val
        self.racing_time = racing_time
        self.position = position

class RacingResultInfo:
    def __init__(self, condition, Gagne_limit, runner_info, result_tables):
        self.condition = condition
        self.Gagne_limit = Gagne_limit
        self.runner_info = runner_info
        self.result_tables = result_tables

class RacingWinningTable:
    def __init__(self, title, row, col, table_val):
        self.title = title
        self.row = row
        self.col = col
        self.table_val = table_val

class FinishRunnerInfo:
    def __init__(self, number, final_time, racing_time):
        self.number = number
        self.final_time = final_time
        self.racing_time = racing_time

def searchInFinishRunnerInfo(final_info, runner_id):
    final_cnt = len(final_info)
    for i in range(0, final_cnt):
        if final_info[i].number==runner_id:
            return i
    return -1

def get_racing_track_info(driver):
    info = []
    soup = BeautifulSoup(driver.page_source, "html.parser")
    try:
        countries = soup.find('div', class_='programme-wrapper').find_all('div', class_='country-bloc')
        for country_item in countries:
            coutry_str = country_item.get_text('|', strip=True)
            racing_table = country_item.find_next_sibling('table', class_='programme')
            racing_items = racing_table.find_all('tr', class_='item')
            for racing_item in racing_items:
                racing_url_str = racing_item.find('td', class_='nom').find('a').get('href')
                racing_track_str = racing_item.find('td', class_='nom').find('h2').get_text('|', strip=True)
                item = RacingTrackInfo(coutry_str, racing_track_str, racing_url_str)
                info.append(item)
    except:
        pass
    return info

def get_racing_general_info(driver, search_type):
    info = []
    soup = BeautifulSoup(driver.page_source, "html.parser")
    try:
        lists = soup.find('div', id='liste-courses').find('table', class_='programme').find_all('tr', class_='item')
        for list_item in lists:
            try:
                racing_name_str = list_item.find('td', class_='nom').find('h3').get_text('|', strip=True)
                racing_url = list_item.find('td', class_='nom').find('a').get('href')
                racing_type = list_item.find('td', class_='type-course').find('span', class_='not-on-mobile').get('title')
                if search_type!="All" and (search_type.lower() not in racing_type.lower()):
                    continue
                racing_starter = list_item.find('td', class_='nb-partants').get_text('|', strip=True)
                racing_distance = list_item.find('td', class_='distance').get_text('|', strip=True)
                racing_runners = list_item.find('td', class_='arrivees').find('span', class_='arrivee').get_text('|', strip=True)
                racing_winner = racing_runners.split('-')[0].strip()
                item = RacingGeneralInfo(racing_name_str, racing_type, racing_starter, racing_distance, racing_winner, racing_runners, racing_url)
                info.append(item)
            except:
                continue
    except:
        pass
    return info

def get_racing_result_info(driver):
    soup = BeautifulSoup(driver.page_source, "html.parser")
    try:
        condition_list = soup.find('p', id='conditions').find('strong').get_text('|', strip=True).split('-')
        condition = ""
        condition_len = len(condition_list)
        for i in range(0, condition_len):
            condition += condition_list[i].strip()
            if i < condition_len-1:
                condition += " - "
        item = soup.find('p', id='conditions').find('strong')
        Gagne_limit = ''
        if item.find_next_sibling('br'):
            Gagne_limit = item.find_next_sibling('br').next_sibling.strip()
    #    time_val = soup.find('div', class_='mCSB_1_container').find('p')
        final_result_header = soup.find('div', id='arriveeTab').find('table').find('thead').find_all('th')
        time_head_id = 0
        racing_time_head_id = 0
        k = 0
        for final_result_header_item in final_result_header:
            ss = final_result_header_item.get_text('|', strip=True)
            if ss=='Time':
                time_head_id = k
            elif ss=='Racing time/km':
                racing_time_head_id = k
            k += 1
        final_result_items = soup.find('div', id='arriveeTab').find('table').find('tbody').findChildren('tr', recursive=False)
        final_result = []
        for final_result_item in final_result_items:
            final_info = final_result_item.findChildren('td', recursive=False)
            final_runner_id = final_info[1].get_text('|', strip=True)
            final_info_len = len(final_info)
            final_time = ""
            final_racing_time = ""
            if final_info_len > 5 and time_head_id > 0 and racing_time_head_id > 0:
                final_time = final_info[time_head_id].get_text('|', strip=True)
                final_racing_time = final_info[racing_time_head_id].get_text('|', strip=True)
            a_final_runner_info = FinishRunnerInfo(final_runner_id, final_time, final_racing_time)
            final_result.append(a_final_runner_info)
        
        runner_tab = driver.find_element_by_id('tab-pari')
        runner_tab.click()
        runner_items = soup.find('table', class_='table-runners').find('tbody').findChildren('tr', role='row')
        runner_info = []
        for runner_item in runner_items:
            try:
                runner_id = runner_item.find('td', class_='numero').find('span', class_='partant').get_text('|', strip=True)
                no_partant = False
                if runner_item.find('td', class_='numero').find('span', class_='non-partant'):
                    no_partant = True
                runner_name = runner_item.find('td', class_='cheval').find('div', class_='first-line').get_text('|', strip=True)
                info_item = runner_item.find('td', class_='information').find('span')
                info = ''
                if info_item:
                    info = info_item.get('title')
                sex_age = runner_item.find('td', class_='sexe-age').get_text('|', strip=True).split('/')
                sex = sex_age[0]
                age = sex_age[1]
                distance = 0
                if runner_item.find('td', class_='distance'):
                    distance = runner_item.find('td', class_='distance').get_text('|', strip=True)
                race_driver_item = runner_item.find('td', class_='driver-entraineur').find('b')
                race_driver = race_driver_item.get_text('|', strip=True)
                race_trainer = race_driver_item.find_next_sibling('br').next_sibling.strip()
                # race_driver_trainer = race_driver_item.get_text('|', strip=True).split('|')
                # if no_partant==False:
                #     race_driver = race_driver_trainer[0].strip()
                #     race_trainer = race_driver_trainer[1].strip()
                # else:
                #     race_trainer = race_driver_trainer[0].strip()
                #     race_driver = ""
                musique = runner_item.find('td', class_='musique').get_text('|', strip=True)
                if no_partant==False:
                    odds = runner_item.find('td', class_='cote').get_text('|', strip=True)
                else:
                    odds = ""
                position = -1
                position = searchInFinishRunnerInfo(final_result, runner_id)
                pos_str = ""
                if position != -1:
                    pos_str = str(position+1)
                    a_runner_info = RacingRunnerInfo(runner_id, runner_name, info, sex, age, distance, race_driver, race_trainer, musique, odds, final_result[position].final_time, final_result[position].racing_time, pos_str)
                else:
                    a_runner_info = RacingRunnerInfo(runner_id, runner_name, info, sex, age, distance, race_driver, race_trainer, musique, odds, "", "", pos_str)
                runner_info.append(a_runner_info)
            except:
                continue
    except:
        return None
    # finish_tab = driver.find_element_by_id('tab-arrivee')
    # finish_tab.click()
    # soup = BeautifulSoup.get(driver.page_source, "html.parser")
    try:
        table_items = soup.find('div', id='rapports').find('div', class_='accordion_body').find_all('div', class_='table')
        table_info = []
        for table_item in table_items:
            table_title = table_item.find('div', class_='table-header').get_text('|', strip=True)
            table_rows = table_item.find('div', class_='table-body').find_all('div', class_='table-row')
            row_count = len(table_rows)
            col_count = 0
            row = 0
            table_value = []
            for table_row in table_rows:
                table_cells = table_row.find_all('div', class_='table-cell')
                col_count = len(table_cells)
                col = 0
                one_row = []
                for table_cell in table_cells:
                    cell_val = table_cell.get_text('|', strip=True)
                    one_row.append(cell_val)
                table_value.append(one_row)
            one_table = RacingWinningTable(table_title, row_count, col_count, table_value)
            table_info.append(one_table)
    except:
        pass
    result_info = RacingResultInfo(condition, Gagne_limit, runner_info, table_info)
    return result_info

def getExcelColStringFromId(col_id):
    alphabet_count = 26
    A_ascii = 65
    div_res = int(col_id/alphabet_count)
    div_mod = col_id%alphabet_count
    first_ch = chr(A_ascii+div_res)
    second_ch = chr(A_ascii+div_mod)
    return first_ch+second_ch

def writeToExcel(worksheet, merge_format, date_format, racing_date, racing_track_info, racing_general_info, url, racing_result_info):
    global line

    worksheet.write_datetime('A'+str(line), racing_date, date_format)
    worksheet.write('B'+str(line), racing_track_info.country)
    worksheet.write('C'+str(line), racing_track_info.race_track)
    worksheet.write('D'+str(line), racing_general_info.racing_name)
    worksheet.write('E'+str(line), racing_general_info.racing_type)
    worksheet.write_url('F'+str(line), url)
    worksheet.write('G'+str(line), racing_result_info.condition)
    worksheet.write('H'+str(line), racing_general_info.starters)
    worksheet.write('I'+str(line), racing_general_info.distance)
    worksheet.write('J'+str(line), racing_general_info.winner)
    worksheet.write('K'+str(line), racing_general_info.runners)
    first_line = line
    runner_info_count = len(racing_result_info.runner_info)
    for i in range(0, runner_info_count):
        a_runner_info = racing_result_info.runner_info[i]
        worksheet.write('L'+str(line), a_runner_info.runner_number)
        worksheet.write('M'+str(line), a_runner_info.runner_name)
        worksheet.write('N'+str(line), a_runner_info.info)
        worksheet.write('O'+str(line), a_runner_info.sex)
        worksheet.write('P'+str(line), a_runner_info.age)
        worksheet.write('Q'+str(line), a_runner_info.distance)
        worksheet.write('R'+str(line), a_runner_info.driver)
        worksheet.write('S'+str(line), a_runner_info.traniner)
        worksheet.write('T'+str(line), a_runner_info.musique)
        worksheet.write('U'+str(line), a_runner_info.odds)
        worksheet.write('V'+str(line), a_runner_info.time_val)
        worksheet.write('W'+str(line), a_runner_info.racing_time)
        worksheet.write('X'+str(line), a_runner_info.position)
        line += 1

    back_line = line
    winning_table_count = len(racing_result_info.result_tables)
    col_id = 0
    for i in range(0, winning_table_count):
        line = first_line
        a_winning_table = racing_result_info.result_tables[i]
        row_count = a_winning_table.row
        col_count = a_winning_table.col
        col_ch1 = getExcelColStringFromId(col_id)+str(line)
        col_ch2 = getExcelColStringFromId(col_id+col_count-1)+str(line)
        worksheet.merge_range(col_ch1+':'+col_ch2, a_winning_table.title, merge_format)
        for row in range(0, row_count):
            line += 1
            for col in range(0, col_count):
                col_ch = getExcelColStringFromId(col_id+col)
                worksheet.write(col_ch+str(line), a_winning_table.table_val[row][col])

        col_id += col_count
    line = back_line

def writeToExcelOpenpy(worksheet, racing_date, racing_track_info, racing_general_info, url, racing_result_info):
    global line

    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")
    date_str = "{:02d}/{:02d}/{:d}".format(racing_date.month, racing_date.day, racing_date.year)
    worksheet['A'+str(line)] = date_str
    worksheet['B'+str(line)] = racing_track_info.country
    worksheet['C'+str(line)] = racing_track_info.race_track
    worksheet['D'+str(line)] = racing_general_info.racing_name
    worksheet['E'+str(line)] = racing_general_info.racing_type
    worksheet['F'+str(line)] = url
    worksheet['G'+str(line)] = racing_result_info.condition
    worksheet['H'+str(line)] = racing_result_info.Gagne_limit
    worksheet['I'+str(line)] = racing_general_info.starters
    worksheet['J'+str(line)] = racing_general_info.distance
    worksheet['K'+str(line)] = racing_general_info.winner
    worksheet['L'+str(line)] = racing_general_info.runners
    first_line = line
    runner_info_count = len(racing_result_info.runner_info)
    for i in range(0, runner_info_count):
        a_runner_info = racing_result_info.runner_info[i]
        worksheet['M'+str(line)] = a_runner_info.runner_number
        worksheet['N'+str(line)] = a_runner_info.runner_name
        worksheet['O'+str(line)] = a_runner_info.info
        worksheet['P'+str(line)] = a_runner_info.sex
        worksheet['Q'+str(line)] = a_runner_info.age
        worksheet['R'+str(line)] = a_runner_info.distance
        worksheet['S'+str(line)] = a_runner_info.driver
        worksheet['T'+str(line)] = a_runner_info.traniner
        worksheet['U'+str(line)] = a_runner_info.musique
        worksheet['V'+str(line)] = a_runner_info.odds
        worksheet['W'+str(line)] = a_runner_info.time_val
        worksheet['X'+str(line)] = a_runner_info.racing_time
        worksheet['Y'+str(line)] = a_runner_info.position
        line += 1
    back_line = line
    winning_table_count = len(racing_result_info.result_tables)
    col_id = 0
    for i in range(0, winning_table_count):
        line = first_line
        a_winning_table = racing_result_info.result_tables[i]
        row_count = a_winning_table.row
        col_count = a_winning_table.col
        col_ch1 = getExcelColStringFromId(col_id)+str(line)
        col_ch2 = getExcelColStringFromId(col_id+col_count-1)+str(line)
        worksheet.merge_cells(col_ch1+':'+col_ch2)
        top_left_cell = worksheet[col_ch1]
        top_left_cell.value = a_winning_table.title
        for j in range(0, col_count):
            col_ch = getExcelColStringFromId(col_id+j)+str(line)
            worksheet[col_ch].border = Border(top=double, left=thin, right=thin, bottom=double)
#        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(0, row_count):
            line += 1
            for col in range(0, col_count):
                col_ch = getExcelColStringFromId(col_id+col)
                worksheet[col_ch+str(line)] = a_winning_table.table_val[row][col]

        col_id += col_count
    line = back_line
    
def start(driver, url, filename, start_date, end_date, racing_type):
    global line
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        make_excel_header_openpy(worksheet)
        # workbook = xlsxwriter.Workbook(filename)
        # worksheet = workbook.add_worksheet()
        # merge_format = workbook.add_format({
        # 'bold': 1,
        # 'border': 1,
        # 'align': 'center',
        # 'valign': 'vcenter',
        # 'fg_color': 'white'})
    #    date_format = workbook.add_format({'num_format': 'mm/dd/yyyy'})
        # make_excel_header(worksheet)
        driver.maximize_window()
        # try:
        #     driver.maximize_window()
        #     driver.get(url)
        #     ttime.sleep(2)
        # except:
        #     try:
        #         driver.get(url)
        #         ttime.sleep(2)
        #     except:
        #         return -1
        curr_date = datetime.datetime.today()
        for year in range(end_date.year, start_date.year-1, -1):
            end_month = 12
            start_month = 0
            if year == end_date.year:
                end_month = end_date.month
                if year==end_date.year and year==start_date.year:
                    start_month = start_date.month-1
            elif year == start_date.year:
                start_month = start_date.month-1
            for month in range(end_month, start_month, -1):
                end_day = DAYS[month-1]
                if month==2 and year%4==0:
                    end_day += 1
                start_day = 0
                if year==end_date.year and month==end_date.month:
                    end_day = end_date.day
                    if year==start_date.year and month==start_date.month:
                        start_day = start_date.day-1
                elif year==start_date.year and month==start_date.month:
                    start_day = start_date.day-1
                for date in range(end_day, start_day, -1):
                    link_url = make_date_url(year, month, date)
    #                link_url = open_page_for_date(driver, url, year, month, date)
                    if link_url == None:
                        continue
                    driver.get(link_url)
                    ttime.sleep(2)
                    racing_date = datetime.date(year, month, date)
                    racing_track_info = get_racing_track_info(driver)
                    for track_info in racing_track_info:
                        link_url = urljoin(url, track_info.url)
                        driver.get(link_url)
                        ttime.sleep(2)
                        racing_general_info = get_racing_general_info(driver, racing_type)
                        for general_info in racing_general_info:
                            link_url = urljoin(url, general_info.url)
                            driver.get(link_url)
                            # if link_url != 'https://www.zeturf.com/en/course/2019-12-31/R2C6-pau-prix-de-bedous/turf':
                            #     continue
                            ttime.sleep(2)
                            racing_result_info = get_racing_result_info(driver)
                            if racing_result_info:
                                writeToExcelOpenpy(worksheet, racing_date, track_info, general_info, link_url, racing_result_info)
                                workbook.save(filename)
                            # if line > 1000:
                            #     workbook.close()
                            #     return 1
        workbook.close()
    except:
        pass
    return 1



if __name__ == "__main__":
    url = "https://www.zeturf.com/en/resultats-et-rapports/archives"
    start(url)