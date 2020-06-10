from bs4 import BeautifulSoup
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time as ttime
import xlsxwriter
import openpyxl
import datetime

def get_change(team, list):
    team_change=[]
    for i in team:
        if i not in list:
            team_change.append(i)
    return team_change

def get_registered_Player(driver, year_string):
    # driver = webdriver.Chrome('./chromedriver')
    driver.get('https://liiga.fi/fi/tilastot/'+year_string+'/runkosarja/pelaajat/')
    ttime.sleep(1)
    registered_Player= {}
    option_team = driver.find_element_by_xpath("//select[@name='team']").find_elements_by_tag_name("option")
    i = 0
    team_count = len(option_team)
    for i in range(1, team_count):
        team_name = option_team[i].text
        option_team[i].click()
        ttime.sleep(1)
        item = []
        registered_Player[team_name]=item
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')

        trs = soup.find('table', id='stats').find('tbody').find_all('tr')
        for tr in trs:
            tds = tr.find_all('td')
            name = tds[1].get_text('|', strip=True) #
            if name=='':
                continue
            name = re.sub('[-=+,#/\?:^$.@*\"※~&%ㆍ!』\\‘|\(\)\[\]\<\>`\'…》]', '', name)
            registered_Player[team_name].append(name)

        option_team = driver.find_element_by_xpath("//select[@name='team']").find_elements_by_tag_name("option")

    return registered_Player
    # driver.close()

def get_running_player(driver, link_url):
    team_home=[]
    team_away=[]
    # driver = webdriver.Chrome('./chromedriver')
    link_url = 'https://liiga.fi'+link_url
    driver.get(link_url)
    html = driver.page_source
    soup_match= BeautifulSoup(html, 'html.parser')
    if soup_match.find('div', class_='team home')==None or soup_match.find('div', class_='team away')==None:
        return None
    home_div=soup_match.find('div', class_='team home').find_all('div', class_='name')
    away_div=soup_match.find('div', class_='team away').find_all('div', class_='name')
    for player in home_div:
        team_home.append(re.sub('[-=+,#/\?:^$.@*\"※~&%ㆍ!』\\‘|\(\)\[\]\<\>`\'…》]', '',player.get_text('|', strip=True)))
    for player in away_div:
        team_away.append(re.sub('[-=+,#/\?:^$.@*\"※~&%ㆍ!』\\‘|\(\)\[\]\<\>`\'…》]', '',player.get_text('|', strip=True)))
    # driver.close()
    return (team_home,team_away)

def get_player_out_new_result(year_string, today, filename, new_out_flg):
    wb = openpyxl.load_workbook(filename)
    idx = wb.sheetnames.index('Liiga_player')
    # sheet1 = wb.active
    wb.remove(wb.worksheets[idx])
    wb.create_sheet('Liiga_player')
    sheet1 = wb.active
    # sheet1 = wb['Sheet']
    sheet1.title = 'Liiga_player'
    sheet1['A1'] = 'Date'
    sheet1['B1'] = 'Home'
    sheet1['C1'] = 'Away'
    sheet1['D1'] = 'Player'
    sheet1['E1'] = 'Team'
    sheet1['F1'] = 'Out/New'
    driver = webdriver.Chrome(executable_path='./chromedrive/chromedriver.exe')
    driver.get('https://liiga.fi/fi/ottelut/'+year_string+'/runkosarja/')
    html = driver.page_source
    soup_games= BeautifulSoup(html, 'html.parser')
    trs_games= soup_games.find('table', id='games').find('tbody').find_all('tr')
    registered_Player=get_registered_Player(driver, year_string)
    for tr in trs_games:
        tds = tr.find_all('td')
        tempdate= tds[1].get_text()
        if tempdate=='':
            date=saved_date
        else :
            date=tempdate
        saved_date=date
        date_string = saved_date.split(' ')[-1].strip()
        d, m, y = [int(x) for x in date_string.split('.')]
        date1 = datetime.date(y, m, d)
        if date1 != today:
            continue
        # if d != 14 or  m != 2 or y != 2020:
        #     continue
        home_away = tds[3].get_text('|', strip=True)
        home = home_away.split('-')[0].strip()
        away = home_away.split('-')[1].strip() 
        link_url = tds[4].find('a', title='Kokoonpanot').get('href')

        running_Player=get_running_player(driver, link_url)
        if running_Player==None:
            continue
        home_Player=running_Player[0]
        away_Player=running_Player[1]
        if new_out_flg==0 or new_out_flg==1:
            home_new=get_change(home_Player,registered_Player[home])
            for player in home_new:
                sheet1.append([date,home,away,player,home,'new'])
                wb.save(filename)
        if new_out_flg==0 or new_out_flg==2:
            home_out=get_change(registered_Player[home],home_Player)
            for player in home_out:
                sheet1.append([date,home,away,player,home,'out'])
                wb.save(filename)

        if new_out_flg==0 or new_out_flg==1:
            away_new=get_change(away_Player,registered_Player[away])
            for player in away_new:
                sheet1.append([date,home,away,player,away,'new'])
                wb.save(filename)
        if new_out_flg==0 or new_out_flg==2:
            away_out=get_change(registered_Player[away],away_Player)
            for player in away_out:
                sheet1.append([date,home,away,player,away,'out'])
                wb.save(filename)
    driver.close()
    wb.save(filename)


   
   
   